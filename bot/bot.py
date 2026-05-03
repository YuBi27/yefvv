"""Telegram quiz bot — ЄФВВ preparation."""
import asyncio
import json
import logging
import os
import random
import re

import redis.asyncio as aioredis
from aiogram import Bot, Dispatcher, F
from aiogram.exceptions import TelegramRetryAfter
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.redis import RedisStorage
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from aiogram.client.default import DefaultBotProperties
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from db import AccessRequest, Base, Question, UserResult

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ["BOT_TOKEN"]
_admin_ids_raw = os.environ.get("ADMIN_IDS", os.environ.get("ADMIN_ID", "0"))
ADMIN_IDS: set[int] = {int(x.strip()) for x in _admin_ids_raw.split(",") if x.strip().isdigit()}
ADMIN_ID = next(iter(ADMIN_IDS), 0)  # primary admin (first in list)
TEACHER_LINK = os.environ.get("TEACHER_LINK", "https://t.me/yevhenia_frolova")
REDIS_URL = os.environ.get("REDIS_URL", "redis://redis:6379")
DATABASE_URL = os.environ.get(
    "DATABASE_URL", "postgresql+asyncpg://bot:bot@postgres:5432/botdb"
)
QUESTIONS_FILE = os.environ.get("QUESTIONS_FILE", "questions.json")
WEBHOOK_DOMAIN = os.environ.get("WEBHOOK_DOMAIN", "")
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")
WEBHOOK_PATH = "/webhook"


# ---------------------------------------------------------------------------
# States
# ---------------------------------------------------------------------------

class UserState(StatesGroup):
    waiting_screenshots = State()
    filling_pib = State()
    filling_study_place = State()
    filling_phone = State()
    filling_email = State()
    filling_instagram = State()
    quiz_in_progress = State()
    quiz_quit_comment = State()  # waiting for quit reason comment
    post_test_score = State()
    broadcast_waiting = State()
    broadcast_filter_set = State()   # admin chose filter, waiting for message
    edit_search = State()
    edit_field = State()
    admin_user_mode = State()   # admin browsing as regular user


# ---------------------------------------------------------------------------
# Flood-safe send helper
# ---------------------------------------------------------------------------

async def safe_send(coro_fn, *args, retries: int = 5, **kwargs):
    """Call coro_fn(*args, **kwargs) with retry on TelegramRetryAfter."""
    for attempt in range(retries):
        try:
            return await coro_fn(*args, **kwargs)
        except TelegramRetryAfter as e:
            wait = e.retry_after + 2
            logger.warning("Flood control: waiting %ds (attempt %d/%d)", wait, attempt + 1, retries)
            await asyncio.sleep(wait)
    return await coro_fn(*args, **kwargs)  # final attempt


# ---------------------------------------------------------------------------
# Keyboards
# ---------------------------------------------------------------------------

def quiz_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📝 Повний тест (140 питань)", callback_data="startquiz:all")],
            [InlineKeyboardButton(text="📚 Пройти за розділами", callback_data="startquiz:sections")],
        ]
    )


SECTIONS_LIST = [
    "КОНСТИТУЦІЙНЕ ПРАВО",
    "АДМІНІСТРАТИВНЕ ПРАВО",
    "ЦИВІЛЬНЕ ПРАВО",
    "ЦИВІЛЬНЕ ПРОЦЕСУАЛЬНЕ ПРАВО",
    "КРИМІНАЛЬНЕ ПРАВО",
    "КРИМІНАЛЬНО-ПРОЦЕСУАЛЬНЕ ПРАВО",
    "МІЖНАРОДНЕ ПУБЛІЧНЕ ПРАВО",
    "МІЖНАРОДНИЙ ЗАХИСТ ПРАВ ЛЮДИНИ",
]

SECTION_EMOJI = {
    "КОНСТИТУЦІЙНЕ ПРАВО": "🏛",
    "АДМІНІСТРАТИВНЕ ПРАВО": "📋",
    "ЦИВІЛЬНЕ ПРАВО": "⚖️",
    "ЦИВІЛЬНЕ ПРОЦЕСУАЛЬНЕ ПРАВО": "📜",
    "КРИМІНАЛЬНЕ ПРАВО": "🔒",
    "КРИМІНАЛЬНО-ПРОЦЕСУАЛЬНЕ ПРАВО": "👨‍⚖️",
    "МІЖНАРОДНЕ ПУБЛІЧНЕ ПРАВО": "🌍",
    "МІЖНАРОДНИЙ ЗАХИСТ ПРАВ ЛЮДИНИ": "🕊",
}


def sections_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(
            text=f"{SECTION_EMOJI.get(s, '📖')} {s.title()}",
            callback_data=f"section:{i}"
        )]
        for i, s in enumerate(SECTIONS_LIST)
    ]
    buttons.append([InlineKeyboardButton(text="◀️ Назад", callback_data="startquiz:back")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def teacher_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👩‍🏫 Зв'язатися з викладачем", url=TEACHER_LINK)]
        ]
    )


def course_keyboard(level: str = "") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👉 2026 рік", callback_data=f"funnel:course4:{level}")],
            [InlineKeyboardButton(text="👉 2027 рік", callback_data=f"funnel:course3:{level}")],
        ]
    )


def interested_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👉 Так, цікаво", callback_data="funnel:interested")],
        ]
    )


def want_more_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👉 Хочу дізнатися більше", callback_data="funnel:want_more")],
        ]
    )


def fill_form_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👉 Хочу заповнити анкету", url="https://forms.gle/X6aMmhxY1Kpe4ZDR9")],
            [InlineKeyboardButton(text="🔄 Почати знову", callback_data="restart_quiz")],
        ]
    )


def restart_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Почати знову", callback_data="restart_quiz")],
        ]
    )


WEBINAR_URL = "https://youtu.be/ClId2Ecu4aQ"
INSTAGRAM_URL = "https://www.instagram.com/yevhenia.frolova"


def approve_keyboard(user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Надати доступ", callback_data=f"access:approve:{user_id}"),
                InlineKeyboardButton(text="❌ Відхилити", callback_data=f"access:reject:{user_id}"),
            ]
        ]
    )


def write_to_user_keyboard(user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✉️ Написати користувачу", url=f"tg://user?id={user_id}")]
        ]
    )


def admin_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📋 Список користувачів які пройшли тест", callback_data="admin:results")],
            [InlineKeyboardButton(text="📝 Заявки на розгляді", callback_data="admin:pending")],
            [InlineKeyboardButton(text="✏️ Редагувати питання", callback_data="admin:edit_questions")],
            [InlineKeyboardButton(text="📣 Розсилка", callback_data="admin:broadcast")],
            [InlineKeyboardButton(text="👤 Режим користувача", callback_data="admin:user_mode")],
        ]
    )


def admin_reply_keyboard() -> ReplyKeyboardMarkup:
    """Persistent bottom keyboard for admin."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📋 Список користувачів"), KeyboardButton(text="📝 Заявки")],
            [KeyboardButton(text="✏️ Питання"), KeyboardButton(text="📣 Розсилка")],
            [KeyboardButton(text="👤 Режим користувача")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def back_to_admin_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👑 Повернутися в панель адміна", callback_data="admin:back_to_admin")],
        ]
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def grade_text(pct: int) -> str:
    if pct >= 90:
        return "🏆 <b>Блискуче!</b> Ти — майбутній правник, якого бояться навіть кодекси!"
    elif pct >= 75:
        return "🎓 <b>Відмінно!</b> Ще трохи — і ти сам будеш складати ці питання."
    elif pct >= 60:
        return "📚 <b>Добре!</b> Фундамент є, але деякі норми ще чекають на тебе вночі."
    elif pct >= 40:
        return "🌱 <b>Є над чим попрацювати.</b> Право — це марафон, а не спринт. Не здавайся!"
    else:
        return "😅 <b>Поки що складно.</b> Але кожен великий юрист колись починав з нуля!"


QUESTION_TIME = 90  # seconds per question


def timer_bar(seconds_left: int, total: int = QUESTION_TIME) -> str:
    """Visual progress bar for timer."""
    filled = round(10 * seconds_left / total)
    bar = "🟩" * filled + "⬜" * (10 - filled)
    if seconds_left > 60:
        emoji = "🟢"
    elif seconds_left > 30:
        emoji = "🟡"
    else:
        emoji = "🔴"
    mins = seconds_left // 60
    secs = seconds_left % 60
    return f"{emoji} {bar} <b>{mins}:{secs:02d}</b>"


# Active timers: chat_id -> asyncio.Task
_timers: dict[int, asyncio.Task] = {}


def cancel_timer(chat_id: int):
    task = _timers.pop(chat_id, None)
    if task and not task.done():
        # Don't cancel if this is the current running task calling us
        current = asyncio.current_task()
        if task is not current:
            task.cancel()


async def run_timer(
    bot: Bot,
    chat_id: int,
    timer_msg_id: int,
    q_id: int,
    q_num: int,
    q_total: int,
    state: FSMContext,
    session_factory,
):
    """Countdown timer — updates message every 10s, auto-skips on timeout."""
    try:
        for remaining in range(QUESTION_TIME - 10, -1, -10):
            await asyncio.sleep(10)
            # Check if question is still active
            data = await state.get_data()
            if not data or data.get("current", -1) != q_num - 1:
                return  # user already answered
            bar = timer_bar(remaining)
            try:
                await bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=timer_msg_id,
                    text=f"⏱ <b>Питання {q_num}/{q_total}</b>\n{bar}",
                    parse_mode="HTML",
                )
            except Exception:
                pass
            if remaining == 0:
                break

        # Time's up — check still on same question
        data = await state.get_data()
        if not data or data.get("current", -1) != q_num - 1:
            return

        q_msg_id_saved = data.get("q_msg_id", 0)

        # Show timeout notification
        timeout_msg = await safe_send(
            bot.send_message,
            chat_id,
            "⌛ <b>Час вийшов!</b> Питання пропущено — відповідь зараховано як неправильну.",
            parse_mode="HTML",
            protect_content=True,
        )

        # Remove answer buttons from question (keep text visible)
        if q_msg_id_saved:
            try:
                await bot.edit_message_reply_markup(
                    chat_id=chat_id, message_id=q_msg_id_saved, reply_markup=None
                )
            except Exception:
                pass

        # Schedule deletion of question, timer and timeout msg after 24h
        msgs_to_delete = [m for m in [q_msg_id_saved, timer_msg_id,
                                       timeout_msg.message_id if timeout_msg else 0] if m]
        if msgs_to_delete:
            asyncio.create_task(_delete_messages_after(bot, chat_id, msgs_to_delete, delay=86400))

        score = data.get("score", 0)
        current = data.get("current", 0)
        await state.update_data(score=score, current=current + 1)
        await asyncio.sleep(1.5)

        # Remove self from timers BEFORE calling send_question (which creates a new timer)
        _timers.pop(chat_id, None)

        async with session_factory() as session:
            await send_question(bot, chat_id, state, session, session_factory)

    except asyncio.CancelledError:
        pass
    finally:
        # Only remove if still pointing to this task (don't remove a newer timer)
        pass

async def _delete_messages_after(bot: Bot, chat_id: int, message_ids: list[int], delay: int = 86400):
    """Delete messages after a delay (default 24h = 86400s)."""
    await asyncio.sleep(delay)
    for msg_id in message_ids:
        try:
            await bot.delete_message(chat_id=chat_id, message_id=msg_id)
        except Exception:
            pass  # already deleted or too old (Telegram allows deletion up to 48h)


async def send_question(bot: Bot, chat_id: int, state: FSMContext, session: AsyncSession, session_factory=None):
    data = await state.get_data()
    q_ids: list[int] = data["q_ids"]
    current: int = data["current"]

    if current >= len(q_ids):
        score: int = data["score"]
        total = len(q_ids)
        pct = score * 100 // total
        user_id: int = data["user_id"]
        username: str = data.get("username", "")
        section_label = data.get("section", "")
        is_admin: bool = data.get("is_admin", False)

        result_obj = UserResult(
            user_id=user_id,
            username=username,
            score=score,
            total=total,
            section=section_label,
            stopped_at=0,
            completed=True,
        )
        session.add(result_obj)
        await session.commit()
        await state.clear()

        # If admin in user mode — show result + back button, skip funnel
        if is_admin:
            await safe_send(
                bot.send_message,
                chat_id,
                f"🏁 <b>Тест завершено!</b>"
                + (f" [{section_label}]" if section_label else "")
                + f"\n\nРезультат: <b>{score}/{total}</b> ({pct}%)\n\n"
                "👑 Ти проходив тест у режимі користувача.",
                parse_mode="HTML",
                reply_markup=back_to_admin_keyboard(),
            )
            return

        # Notify user — post-test funnel message #1
        await safe_send(
            bot.send_message,
            chat_id,
            f"🏁 <b>Тест завершено!</b>"
            + (f" [{section_label}]" if section_label else "")
            + f"\n\nТвій результат: <b>{score}/{total}</b> ({pct}%)\n\n"
            "Вітаю! Напишіть свій результат нижче — безкоштовно дам зворотній зв'язок, що це означає.\n\n"
            "Або обери діапазон:\n"
            "• До 35\n• 35–50\n• 51–80\n• 80–140",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="До 35", callback_data=f"score_range:low:{user_id}")],
                [InlineKeyboardButton(text="35–50", callback_data=f"score_range:satisf:{user_id}")],
                [InlineKeyboardButton(text="51–80", callback_data=f"score_range:mid:{user_id}")],
                [InlineKeyboardButton(text="80–140", callback_data=f"score_range:high:{user_id}")],
            ]),
        )

        # Notify all admins
        if ADMIN_IDS:
            req = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == user_id)
            )).scalar()
            pib = req.pib if req and req.pib else "—"
            phone = req.phone if req and req.phone else "—"
            email = req.email if req and req.email else "—"
            study = req.study_place if req and req.study_place else "—"
            course_val = req.course if req and req.course else "—"
            insta = req.instagram if req and req.instagram else "—"
            uname = f"@{username}" if username else "—"
            admin_text = (
                f"📊 <b>Новий результат тесту</b>\n\n"
                f"👤 ПІБ: {pib}\n"
                f"🔗 Telegram: {uname}\n"
                f"📞 Телефон: {phone}\n"
                f"📧 Email: {email}\n"
                f"🏫 Навчання: {study}, {course_val}\n"
                f"📸 Instagram: {insta}\n"
                f"🆔 ID: <code>{user_id}</code>\n\n"
                + (f"📚 Розділ: {section_label}\n" if section_label else "")
                + f"🏆 Результат: <b>{score}/{total}</b> ({pct}%)"
            )
            for admin_id in ADMIN_IDS:
                await safe_send(
                    bot.send_message,
                    admin_id, admin_text,
                    parse_mode="HTML",
                    reply_markup=write_to_user_keyboard(user_id),
                )
        return

    q_id = q_ids[current]
    result = await session.execute(select(Question).where(Question.id == q_id))
    q = result.scalar_one()

    question_text = re.sub(r"\s+\d+\s*$", "", q.question).strip()
    options_text = "\n\n".join(
        f"<b>{chr(65+i)})</b> {opt.strip()}" for i, opt in enumerate(q.options)
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=chr(65+i), callback_data=f"ans:{q_id}:{i}")
         for i in range(len(q.options))],
        [InlineKeyboardButton(text="🚪 Завершити тест", callback_data="quiz:quit")],
    ])

    # Cancel any existing timer for this chat
    cancel_timer(chat_id)

    # Send question message
    q_msg = await safe_send(
        bot.send_message,
        chat_id,
        f"❓ <b>Питання {current + 1}/{len(q_ids)}</b>\n\n"
        f"{question_text}\n\n"
        f"——————————————\n"
        f"{options_text}",
        parse_mode="HTML",
        reply_markup=keyboard,
        protect_content=True,
    )

    # Send timer message right after
    timer_msg = await safe_send(
        bot.send_message,
        chat_id,
        f"⏱ <b>Питання {current + 1}/{len(q_ids)}</b>\n{timer_bar(QUESTION_TIME)}",
        parse_mode="HTML",
        protect_content=True,
    )

    # Save message IDs in state so handle_answer can delete them
    if q_msg and timer_msg:
        await state.update_data(
            q_msg_id=q_msg.message_id,
            timer_msg_id=timer_msg.message_id,
        )

    if timer_msg:
        task = asyncio.create_task(run_timer(
            bot=bot,
            chat_id=chat_id,
            timer_msg_id=timer_msg.message_id,
            q_id=q_id,
            q_num=current + 1,
            q_total=len(q_ids),
            state=state,
            session_factory=session_factory,
        ))
        _timers[chat_id] = task


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main():
    redis = aioredis.from_url(REDIS_URL)
    storage = RedisStorage(redis)

    engine = create_async_engine(DATABASE_URL, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    session_factory = async_sessionmaker(engine, expire_on_commit=False)

    # Seed questions
    async with session_factory() as session:
        if not (await session.execute(select(Question).limit(1))).scalar():
            with open(QUESTIONS_FILE, encoding="utf-8") as f:
                qs = json.load(f)
            for q in qs:
                session.add(Question(
                    question=q["question"],
                    options=q["options"],
                    correct=q["correct"],
                    section=q.get("section", ""),
                ))
            await session.commit()
            logger.info("Seeded %d questions", len(qs))

    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="HTML", protect_content=True,))
    dp = Dispatcher(storage=storage)

    # -----------------------------------------------------------------------
    # /start
    # -----------------------------------------------------------------------
    @dp.message(Command("start"))
    async def cmd_start(message: Message, state: FSMContext):
        # Admin shortcut — but check if in user mode first
        if message.from_user.id in ADMIN_IDS:
            current_state = await state.get_state()
            if current_state == UserState.admin_user_mode:
                # Already in user mode — show quiz menu with back button
                await message.answer(
                    "👤 Ти в режимі користувача. Обери формат тесту:",
                    reply_markup=quiz_menu_keyboard(),
                )
                await message.answer(
                    "↩️ Або повернись в панель адміна:",
                    reply_markup=back_to_admin_keyboard(),
                )
                return
            # Normal admin start
            await safe_send(
                message.answer,
                "👑 Панель адміна — обери дію:",
                reply_markup=admin_reply_keyboard(),
            )
            return

        async with session_factory() as session:
            req = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == message.from_user.id)
            )).scalar()

        if req and req.approved:
            # Check if profile is filled
            if req.pib:
                await message.answer("👋 Вітаю! Обери формат тесту:", reply_markup=quiz_menu_keyboard())
            else:
                await state.set_state(UserState.filling_pib)
                await message.answer(
                    "🎉 Доступ підтверджено!\n\n"
                    "Перед початком тесту заповни свій профіль.\n\n"
                    "✏️ Введи своє <b>ПІБ</b> (Прізвище Ім'я По-батькові):",
                    parse_mode="HTML",
                )
        elif req and req.status == "pending":
            await message.answer("⏳ Твоя заявка вже на розгляді. Очікуй підтвердження від адміністратора.")
        elif req and req.status == "rejected":
            await message.answer(
                "❌ На жаль, твою заявку відхилено.\n"
                "Якщо вважаєш це помилкою — зв'яжись з викладачем.",
                reply_markup=teacher_keyboard(),
            )
        else:
            await state.set_state(UserState.waiting_screenshots)
            await message.answer(
                "🔶 Вітаю! Це бот Євгенії Фролової — Мами ЄФВВ та ЄДКІ.\n\n"
                "Щоб отримати доступ до <b>АВТОРСЬКОЇ СИМУЛЯЦІЇ</b> тесту ЄФВВ з права, "
                "якої <b>НЕМАЄ У ВІДКРИТОМУ ДОСТУПІ</b>, виконайте умови:\n\n"
                "🤎 1. Підпишіться на Instagram <a href='https://www.instagram.com/yevhenia.frolova'>@yevhenia.frolova</a> — якщо ви ще не підписані.\n\n"
                "🤎 2. Зробіть репост <b>ЦІЄЇ публікації</b> [вставимо посилання у понеділок] у свої stories з відміткою <a href='https://www.instagram.com/yevhenia.frolova'>@yevhenia.frolova</a>\n\n"
                "🤎 3. Сторіс має бути активна <b>24 години</b>.\n\n"
                "🤎 4. Сторінка на цей час має бути <b>ВІДКРИТА</b>.\n\n"
                "🔶 Після виконання умов:\n"
                "👉 надішліть <b>СКРІНШОТИ</b> підписки та репосту <b>ПРЯМО СЮДИ</b>.\n"
                "👉 адміністратор перевірить і надасть <b>ДОСТУП ДО СИМУЛЯЦІЇ</b>.\n\n"
                "Чекаю на тебе 🤎",
                parse_mode="HTML",
                disable_web_page_preview=True,
            )

    # -----------------------------------------------------------------------
    # -----------------------------------------------------------------------
    # Profile filling flow (after approval)
    # -----------------------------------------------------------------------
    @dp.message(UserState.filling_pib)
    async def fill_pib(message: Message, state: FSMContext):
        pib = message.text.strip() if message.text else ""
        parts = pib.split()
        if len(parts) < 2 or any(len(p) < 2 for p in parts) or not all(c.isalpha() or c in " '-ʼ" for c in pib):
            await message.answer(
                "⚠️ Схоже, ПІБ введено некоректно.\n\n"
                "Будь ласка, введи повне <b>Прізвище Ім'я По-батькові</b> (мінімум 2 слова, лише літери).\n\n"
                "Наприклад: <i>Іваненко Іван Іванович</i>",
                parse_mode="HTML",
            )
            return
        await state.update_data(pib=pib)
        await state.set_state(UserState.filling_study_place)
        await message.answer(
            "🏫 Введи своє <b>місце навчання та курс</b>\n\n"
            "Наприклад: <i>КНУ ім. Шевченка, 4 курс</i>",
            parse_mode="HTML",
        )

    @dp.message(UserState.filling_study_place)
    async def fill_study_place(message: Message, state: FSMContext):
        text = message.text.strip() if message.text else ""
        if len(text) < 3:
            await message.answer("⚠️ Введи місце навчання та курс (наприклад: КНУ, 4 курс).")
            return
        # Try to extract course number
        course = ""
        for word in text.split():
            if word.isdigit() and 1 <= int(word) <= 6:
                course = f"{word} курс"
                break
        await state.update_data(study_place=text, course=course)
        await state.set_state(UserState.filling_phone)
        await message.answer("📞 Введи свій <b>номер телефону</b>:", parse_mode="HTML")

    @dp.message(UserState.filling_phone)
    async def fill_phone(message: Message, state: FSMContext):
        phone = message.text.strip() if message.text else ""
        digits = "".join(c for c in phone if c.isdigit())
        if len(digits) < 9 or len(digits) > 15:
            await message.answer(
                "⚠️ Номер телефону введено некоректно.\n\n"
                "Введи номер у форматі <b>+380XXXXXXXXX</b> або <b>0XXXXXXXXX</b>.",
                parse_mode="HTML",
            )
            return
        await state.update_data(phone=phone)
        await state.set_state(UserState.filling_email)
        await message.answer("📧 Введи свою <b>електронну пошту</b>:", parse_mode="HTML")

    @dp.message(UserState.filling_email)
    async def fill_email(message: Message, state: FSMContext):
        email = message.text.strip() if message.text else ""
        if "@" not in email or "." not in email.split("@")[-1] or len(email) < 5:
            await message.answer(
                "⚠️ Email введено некоректно.\n\n"
                "Введи дійсну адресу, наприклад: <i>example@gmail.com</i>",
                parse_mode="HTML",
            )
            return
        await state.update_data(email=email)
        await state.set_state(UserState.filling_instagram)
        await message.answer(
            "📸 Введи посилання на свій профіль в <b>Instagram</b>\n\n"
            "Наприклад: <i>https://instagram.com/username</i> або просто <i>@username</i>",
            parse_mode="HTML",
        )

    @dp.message(UserState.filling_instagram)
    async def fill_instagram(message: Message, state: FSMContext):
        instagram = message.text.strip() if message.text else ""
        if len(instagram) < 2:
            await message.answer("⚠️ Введи посилання або нікнейм Instagram.")
            return

        data = await state.get_data()
        pib = data["pib"]
        study_place = data["study_place"]
        course = data["course"]
        phone = data["phone"]
        email = data["email"]

        async with session_factory() as session:
            req = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == message.from_user.id)
            )).scalar()
            if req:
                req.pib = pib
                req.study_place = study_place
                req.course = course
                req.phone = phone
                req.email = email
                req.instagram = instagram
                await session.commit()

        await state.clear()
        await message.answer(
            f"✅ <b>Профіль збережено!</b>\n\n"
            f"👤 ПІБ: {pib}\n"
            f"🏫 Навчання: {study_place}\n"
            f"📞 Телефон: {phone}\n"
            f"📧 Email: {email}\n"
            f"📸 Instagram: {instagram}\n\n"
            "Ти зараз пройдеш симуляцію ЄФВВ — це формат, який показує твій реальний рівень підготовки.\n\n"
            "Кількість запитань — 140, як на реальному ЄФВВ.\n\n"
            "<b>Важливо:</b>\n"
            "— не гугли\n"
            "— не виправляй\n"
            "— контролюй час (на реальному іспиті: 1 питання — 1 хв 30 с)\n\n"
            "Інакше — ти обманеш тільки себе.\n\n"
            "Головне — за допомогою цього тесту перевірити себе, свої знання, витривалість та вміння концентрувати увагу.\n\n"
            "<b>НАПИШИ ПОТІМ СВІЙ РЕЗУЛЬТАТ — Я ДАМ БЕЗКОШТОВНО ЗВОРОТНІЙ ЗВ'ЯЗОК, ЩО ЦЕ ОЗНАЧАЄ.</b>\n\n"
            "P.S. Випускники з року в рік відзначають, що ця симуляція допомогла їм вчасно виявити прогалини, підтягнути слабкі теми і не втратити «гарантовані» бали на іспиті.\n\n"
            "👇 Обери формат тесту:",
            parse_mode="HTML",
            reply_markup=quiz_menu_keyboard(),
        )

    # -----------------------------------------------------------------------
    # Screenshots from user
    # -----------------------------------------------------------------------
    @dp.message(UserState.waiting_screenshots, F.photo | F.document)
    async def receive_screenshots(message: Message, state: FSMContext):
        user = message.from_user
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        username_str = f"@{user.username}" if user.username else "немає"

        async with session_factory() as session:
            existing = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == user.id)
            )).scalar()

            if existing:
                if existing.status == "pending":
                    await message.answer("⏳ Твоя заявка вже надіслана. Очікуй відповіді.")
                    return
                elif existing.approved:
                    await message.answer("✅ У тебе вже є доступ! Напиши /start")
                    await state.clear()
                    return
                existing.status = "pending"
                existing.approved = False
                await session.commit()
            else:
                session.add(AccessRequest(
                    user_id=user.id,
                    username=user.username or "",
                    full_name=full_name,
                    status="pending",
                ))
                await session.commit()

        await message.answer("✅ Скріншоти отримано! Адміністратор перевірить і надасть доступ найближчим часом.")
        await state.clear()

        if ADMIN_IDS:
            for admin_id in ADMIN_IDS:
                await bot.forward_message(admin_id, message.chat.id, message.message_id)
                await bot.send_message(
                    admin_id,
                    f"📋 <b>Нова заявка на доступ</b>\n\n"
                    f"👤 {full_name}\n"
                    f"🔗 {username_str}\n"
                    f"🆔 <code>{user.id}</code>",
                    parse_mode="HTML",
                    reply_markup=approve_keyboard(user.id),
                )

    @dp.message(UserState.waiting_screenshots)
    async def waiting_wrong_type(message: Message):
        await message.answer("📸 Будь ласка, надішли скріншоти (фото або файл).")

    # -----------------------------------------------------------------------
    # Admin: approve / reject callback
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data.startswith("access:"))
    async def handle_access_decision(callback: CallbackQuery):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔ Тільки для адміна.", show_alert=True)
            return

        _, action, user_id_str = callback.data.split(":")
        target_user_id = int(user_id_str)

        async with session_factory() as session:
            req = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == target_user_id)
            )).scalar()

            if not req:
                await callback.answer("Заявку не знайдено.", show_alert=True)
                return

            if action == "approve":
                req.status = "approved"
                req.approved = True
                await session.commit()
                await callback.message.edit_reply_markup(reply_markup=None)
                await callback.answer("✅ Доступ надано!")
                await bot.send_message(
                    target_user_id,
                    "🎉 <b>Доступ надано!</b>\n\n"
                    "Перед початком тесту заповни свій профіль.\n\n"
                    "✏️ Введи своє <b>ПІБ</b> (Прізвище Ім'я По-батькові):",
                    parse_mode="HTML",
                )
                from aiogram.fsm.storage.base import StorageKey
                user_state = FSMContext(
                    storage=storage,
                    key=StorageKey(bot_id=bot.id, chat_id=target_user_id, user_id=target_user_id),
                )
                await user_state.set_state(UserState.filling_pib)

            elif action == "reject":
                req.status = "rejected"
                req.approved = False
                await session.commit()
                await callback.message.edit_reply_markup(reply_markup=None)
                await callback.answer("❌ Заявку відхилено.")
                await bot.send_message(
                    target_user_id,
                    "😔 На жаль, твою заявку відхилено.\n"
                    "Переконайся, що виконав всі умови, і спробуй ще раз (/start).",
                )

    # -----------------------------------------------------------------------
    # Admin panel callbacks
    # -----------------------------------------------------------------------
    @dp.message(Command("admin"))
    async def cmd_admin(message: Message):
        if message.from_user.id not in ADMIN_IDS:
            return
        await safe_send(
            message.answer,
            "👑 Панель адміна — обери дію:",
            reply_markup=admin_reply_keyboard(),
        )

    @dp.message(F.text == "🏠 Меню")
    async def admin_home(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        # Exit user mode if active
        await state.clear()
        await safe_send(
            message.answer,
            "👑 Панель адміна — обери дію:",
            reply_markup=admin_reply_keyboard(),
        )

    @dp.message(F.text == "👤 Режим користувача")
    async def admin_user_mode_btn(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        await state.set_state(UserState.admin_user_mode)
        await message.answer(
            "👤 <b>Режим користувача активовано</b>\n\n"
            "Тепер ти бачиш бот очима користувача.\n"
            "Обери формат тесту 👇",
            parse_mode="HTML",
            reply_markup=quiz_menu_keyboard(),
        )
        await message.answer(
            "↩️ Повернутися в панель адміна:",
            reply_markup=back_to_admin_keyboard(),
        )

    @dp.callback_query(F.data == "admin:user_mode")
    async def admin_user_mode_cb(callback: CallbackQuery, state: FSMContext):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        await callback.answer()
        await state.set_state(UserState.admin_user_mode)
        await callback.message.answer(
            "👤 <b>Режим користувача активовано</b>\n\n"
            "Тепер ти бачиш бот очима користувача.\n"
            "Обери формат тесту 👇",
            parse_mode="HTML",
            reply_markup=quiz_menu_keyboard(),
        )
        await callback.message.answer(
            "↩️ Повернутися в панель адміна:",
            reply_markup=back_to_admin_keyboard(),
        )

    @dp.callback_query(F.data == "admin:back_to_admin")
    async def back_to_admin_cb(callback: CallbackQuery, state: FSMContext):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        await callback.answer()
        await state.clear()
        await callback.message.edit_reply_markup(reply_markup=None)
        await safe_send(
            callback.message.answer,
            "👑 Панель адміна — обери дію:",
            reply_markup=admin_reply_keyboard(),
        )

    @dp.message(F.text == "📋 Список користувачів")
    async def admin_users_btn(message: Message):
        if message.from_user.id not in ADMIN_IDS:
            return
        await _show_results(message)

    @dp.message(F.text == "📝 Заявки")
    async def admin_pending_btn(message: Message):
        if message.from_user.id not in ADMIN_IDS:
            return
        await _show_pending(message)

    @dp.message(F.text == "📣 Розсилка")
    async def admin_broadcast_btn(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        await _start_broadcast(message, state)

    @dp.callback_query(F.data == "admin:pending")
    async def admin_pending(callback: CallbackQuery):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        await callback.answer()
        await _show_pending(callback.message)

    async def _show_pending(target: Message):
        async with session_factory() as session:
            rows = (await session.execute(
                select(AccessRequest).where(AccessRequest.status == "pending")
            )).scalars().all()
        if not rows:
            await target.answer("Немає заявок на розгляді.")
            return
        for req in rows:
            uname = f"@{req.username}" if req.username else "немає"
            await target.answer(
                f"👤 {req.full_name}\n🔗 {uname}\n🆔 <code>{req.user_id}</code>",
                parse_mode="HTML",
                reply_markup=approve_keyboard(req.user_id),
            )

    PAGE_SIZE = 5  # results per page

    def results_page_keyboard(page: int, total_pages: int, total: int) -> InlineKeyboardMarkup:
        buttons = []
        nav = []
        if page > 0:
            nav.append(InlineKeyboardButton(text="◀️", callback_data=f"results_page:{page-1}"))
        nav.append(InlineKeyboardButton(
            text=f"📄 {page+1}/{total_pages}  ({total} записів)",
            callback_data="results_noop"
        ))
        if page < total_pages - 1:
            nav.append(InlineKeyboardButton(text="▶️", callback_data=f"results_page:{page+1}"))
        buttons.append(nav)
        buttons.append([
            InlineKeyboardButton(text="🔄 Оновити", callback_data="results_page:0"),
        ])
        buttons.append([
            InlineKeyboardButton(text="📥 Експорт CSV", callback_data="results_export:csv"),
            InlineKeyboardButton(text="📊 Експорт Excel", callback_data="results_export:xlsx"),
        ])
        return InlineKeyboardMarkup(inline_keyboard=buttons)

    async def _fetch_all_results():
        """Fetch all results within 7 days with profiles."""
        from datetime import datetime, timedelta
        cutoff = datetime.utcnow() - timedelta(days=7)
        async with session_factory() as session:
            all_results = (await session.execute(
                select(UserResult)
                .where(UserResult.created_at >= cutoff)
                .order_by(UserResult.created_at.desc())
            )).scalars().all()
            if not all_results:
                return [], {}
            user_ids = list({r.user_id for r in all_results})
            reqs = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id.in_(user_ids))
            )).scalars().all()
            profiles = {r.user_id: r for r in reqs}
        return all_results, profiles

    async def _get_results_page(page: int) -> tuple[str, int, int]:
        """Returns (text, total_pages, total_count)."""
        all_results, profiles = await _fetch_all_results()
        if not all_results:
            return "За останні 7 днів ніхто не проходив тест.", 0, 0

        total = len(all_results)
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
        page = max(0, min(page, total_pages - 1))
        chunk = all_results[page * PAGE_SIZE:(page + 1) * PAGE_SIZE]

        lines = [f"📊 <b>Список користувачів які пройшли тест</b>\nЗа 7 днів: {total} проходжень\n"]
        for i, r in enumerate(chunk, page * PAGE_SIZE + 1):
            prof = profiles.get(r.user_id)
            pib = prof.pib if prof and prof.pib else "—"
            phone = prof.phone if prof and prof.phone else "—"
            email = prof.email if prof and prof.email else "—"
            study = prof.study_place if prof and prof.study_place else "—"
            course_val = prof.course if prof and prof.course else "—"
            insta = prof.instagram if prof and prof.instagram else "—"
            uname = f"@{r.username}" if r.username else "—"
            pct = r.score * 100 // r.total
            date_str = r.created_at.strftime("%d.%m %H:%M") if r.created_at else "—"

            # Test info
            section_info = r.section if r.section else "Повний тест"
            if r.completed:
                test_status = f"✅ Завершено ({r.score}/{r.total}, {pct}%)"
            else:
                test_status = f"⏸ Зупинився на питанні {r.stopped_at}/{r.total}"

            lines.append(
                f"<b>#{i}</b> 👤 {pib}\n"
                f"   🔗 {uname}  🆔 <code>{r.user_id}</code>\n"
                f"   📞 {phone}  📧 {email}\n"
                f"   🏫 {study} ({course_val})\n"
                f"   📸 {insta}\n"
                f"   📚 {section_info}\n"
                f"   {test_status}  📅 {date_str}"
            )

        return "\n\n".join(lines), total_pages, total

    def _build_csv(all_results, profiles) -> bytes:
        import csv, io
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "№", "ПІБ", "Telegram", "Телефон", "Email",
            "Навчання", "Курс", "Instagram", "Розділ",
            "Статус", "Бал", "Всього", "%", "Зупинився на питанні", "Дата"
        ])
        for i, r in enumerate(all_results, 1):
            prof = profiles.get(r.user_id)
            pct = r.score * 100 // r.total
            status = "Завершено" if r.completed else "Не завершено"
            writer.writerow([
                i,
                prof.pib if prof else "",
                f"@{r.username}" if r.username else "",
                prof.phone if prof else "",
                prof.email if prof else "",
                prof.study_place if prof else "",
                prof.course if prof else "",
                prof.instagram if prof else "",
                r.section or "Повний тест",
                status,
                r.score, r.total, pct,
                r.stopped_at if not r.completed else "",
                r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else "",
            ])
        return buf.getvalue().encode("utf-8-sig")

    def _build_xlsx(all_results, profiles) -> bytes:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Результати"

        headers = [
            "№", "ПІБ", "Telegram", "Телефон", "Email",
            "Навчання", "Курс", "Instagram", "Розділ",
            "Статус", "Бал", "Всього", "%", "Зупинився на питанні", "Дата"
        ]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(all_results, 1):
            prof = profiles.get(r.user_id)
            pct = r.score * 100 // r.total
            status = "Завершено" if r.completed else "Не завершено"
            row = [
                i,
                prof.pib if prof else "",
                f"@{r.username}" if r.username else "",
                prof.phone if prof else "",
                prof.email if prof else "",
                prof.study_place if prof else "",
                prof.course if prof else "",
                prof.instagram if prof else "",
                r.section or "Повний тест",
                status,
                r.score, r.total, pct,
                r.stopped_at if not r.completed else "",
                r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else "",
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=i + 1, column=col, value=val)
            # Color row red if not completed
            if not r.completed:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i + 1, column=col).fill = PatternFill("solid", fgColor="FFE0E0")

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @dp.callback_query(F.data.startswith("results_export:"))
    async def results_export_cb(callback: CallbackQuery):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        fmt = callback.data.split(":")[1]
        await callback.answer("⏳ Генерую файл...")

        all_results, profiles = await _fetch_all_results()
        if not all_results:
            await callback.message.answer("Немає даних для експорту.")
            return

        from aiogram.types import BufferedInputFile
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M")

        if fmt == "csv":
            data = _build_csv(all_results, profiles)
            file = BufferedInputFile(data, filename=f"results_{ts}.csv")
            await callback.message.answer_document(file, caption="📥 Експорт CSV")
        else:
            data = _build_xlsx(all_results, profiles)
            file = BufferedInputFile(data, filename=f"results_{ts}.xlsx")
            await callback.message.answer_document(file, caption="📊 Експорт Excel")

    @dp.callback_query(F.data == "admin:results")
    async def admin_results(callback: CallbackQuery):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        await callback.answer()
        text, total_pages, total = await _get_results_page(0)
        if total == 0:
            await callback.message.answer(text)
            return
        await callback.message.answer(
            text,
            parse_mode="HTML",
            reply_markup=results_page_keyboard(0, total_pages, total),
        )

    @dp.callback_query(F.data.startswith("results_page:"))
    async def results_page_cb(callback: CallbackQuery):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        page = int(callback.data.split(":")[1])
        text, total_pages, total = await _get_results_page(page)
        if total == 0:
            await callback.answer("Немає даних.", show_alert=True)
            return
        try:
            await callback.message.edit_text(
                text,
                parse_mode="HTML",
                reply_markup=results_page_keyboard(page, total_pages, total),
            )
        except Exception:
            pass
        await callback.answer()

    @dp.callback_query(F.data == "results_noop")
    async def results_noop(callback: CallbackQuery):
        await callback.answer()

    async def _show_results(target: Message):
        text, total_pages, total = await _get_results_page(0)
        if total == 0:
            await target.answer(text)
            return
        await target.answer(
            text,
            parse_mode="HTML",
            reply_markup=results_page_keyboard(0, total_pages, total),
        )

    @dp.callback_query(F.data == "admin:broadcast")
    async def admin_broadcast_cb(callback: CallbackQuery, state: FSMContext):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        await callback.answer()
        await _start_broadcast(callback.message, state)

    @dp.message(Command("cancel"))
    async def cmd_cancel(message: Message, state: FSMContext):
        current = await state.get_state()
        if current:
            await state.clear()
            await message.answer("❌ Скасовано.")

    # -----------------------------------------------------------------------
    # Admin: reply keyboard button for questions
    # -----------------------------------------------------------------------
    @dp.message(F.text == "✏️ Питання")
    async def admin_edit_btn(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        await _start_edit_questions(message, state)

    @dp.callback_query(F.data == "admin:edit_questions")
    async def admin_edit_cb(callback: CallbackQuery, state: FSMContext):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        await callback.answer()
        await _start_edit_questions(callback.message, state)

    async def _start_edit_questions(target: Message, state: FSMContext):
        await state.set_state(UserState.edit_search)
        await target.answer(
            "✏️ <b>Редактор питань</b>\n\n"
            "Введи <b>номер питання</b> (1–140) яке хочеш редагувати:",
            parse_mode="HTML",
        )

    @dp.message(UserState.edit_search)
    async def edit_search_handler(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        text = message.text.strip() if message.text else ""
        if not text.isdigit() or not (1 <= int(text) <= 140):
            await message.answer("⚠️ Введи число від 1 до 140.")
            return
        q_num = int(text)
        async with session_factory() as session:
            q = (await session.execute(
                select(Question).order_by(Question.id).offset(q_num - 1).limit(1)
            )).scalar()
        if not q:
            await message.answer("❌ Питання не знайдено.")
            return

        await state.update_data(edit_q_id=q.id, edit_q_num=q_num)
        await state.set_state(None)  # clear state, use callback flow

        opts = "\n".join(f"  <b>{chr(65+i)})</b> {o}" for i, o in enumerate(q.options))
        correct_letter = chr(65 + q.correct)
        await message.answer(
            f"📋 <b>Питання {q_num}</b> [ID: {q.id}]\n"
            f"📚 Розділ: {q.section}\n\n"
            f"<b>{q.question}</b>\n\n"
            f"{opts}\n\n"
            f"✅ Правильна відповідь: <b>{correct_letter}</b>\n\n"
            "Що редагуємо?",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📝 Текст питання", callback_data=f"eq:question:{q.id}")],
                [InlineKeyboardButton(text="🅰️ Варіант A", callback_data=f"eq:opt0:{q.id}")],
                [InlineKeyboardButton(text="🅱️ Варіант B", callback_data=f"eq:opt1:{q.id}")],
                [InlineKeyboardButton(text="🅲 Варіант C", callback_data=f"eq:opt2:{q.id}")],
                [InlineKeyboardButton(text="🅳 Варіант D", callback_data=f"eq:opt3:{q.id}")],
                [InlineKeyboardButton(text="✅ Правильна відповідь (A/B/C/D)", callback_data=f"eq:correct:{q.id}")],
                [InlineKeyboardButton(text="🔍 Інше питання", callback_data="eq:search")],
            ]),
        )

    @dp.callback_query(F.data.startswith("eq:"))
    async def edit_question_cb(callback: CallbackQuery, state: FSMContext):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        parts = callback.data.split(":")
        action = parts[1]
        await callback.answer()

        if action == "search":
            await callback.message.edit_reply_markup(reply_markup=None)
            await state.set_state(UserState.edit_search)
            await callback.message.answer("Введи номер питання (1–140):")
            return

        q_id = int(parts[2])
        await state.update_data(edit_q_id=q_id, edit_field=action)
        await state.set_state(UserState.edit_field)

        prompts = {
            "question": "📝 Введи новий текст питання:",
            "opt0": "🅰️ Введи новий текст варіанту <b>A</b>:",
            "opt1": "🅱️ Введи новий текст варіанту <b>B</b>:",
            "opt2": "🅲 Введи новий текст варіанту <b>C</b>:",
            "opt3": "🅳 Введи новий текст варіанту <b>D</b>:",
            "correct": "✅ Введи правильну відповідь: <b>A</b>, <b>B</b>, <b>C</b> або <b>D</b>:",
        }
        await callback.message.answer(prompts.get(action, "Введи нове значення:"), parse_mode="HTML")

    @dp.message(UserState.edit_field)
    async def edit_field_handler(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        data = await state.get_data()
        q_id = data["edit_q_id"]
        field = data["edit_field"]
        new_val = message.text.strip() if message.text else ""

        if not new_val:
            await message.answer("⚠️ Порожнє значення. Спробуй ще раз.")
            return

        async with session_factory() as session:
            q = (await session.execute(select(Question).where(Question.id == q_id))).scalar()
            if not q:
                await message.answer("❌ Питання не знайдено.")
                await state.clear()
                return

            if field == "question":
                q.question = new_val
            elif field.startswith("opt"):
                idx = int(field[3])
                opts = list(q.options)
                opts[idx] = new_val
                q.options = opts
            elif field == "correct":
                letter = new_val.upper().strip()
                if letter not in ("A", "B", "C", "D"):
                    await message.answer("⚠️ Введи A, B, C або D.")
                    return
                q.correct = ord(letter) - ord("A")

            await session.commit()

            # Show updated question
            opts_text = "\n".join(f"  <b>{chr(65+i)})</b> {o}" for i, o in enumerate(q.options))
            correct_letter = chr(65 + q.correct)

        await state.clear()
        await message.answer(
            f"✅ <b>Збережено!</b>\n\n"
            f"<b>{q.question}</b>\n\n"
            f"{opts_text}\n\n"
            f"✅ Правильна відповідь: <b>{correct_letter}</b>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✏️ Редагувати ще", callback_data=f"eq:question:{q_id}")],
                [InlineKeyboardButton(text="🔍 Інше питання", callback_data="eq:search")],
            ]),
        )

    # -----------------------------------------------------------------------
    # Broadcast — filter keyboard and logic
    # -----------------------------------------------------------------------
    def broadcast_filter_keyboard() -> InlineKeyboardMarkup:
        return InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="👥 Всі користувачі", callback_data="bcast_filter:all")],
            [InlineKeyboardButton(text="✅ Пройшли тест до кінця", callback_data="bcast_filter:completed")],
            [InlineKeyboardButton(text="⏸ Не завершили тест", callback_data="bcast_filter:not_completed")],
            [InlineKeyboardButton(text="🔴 Погано (до 35 балів)", callback_data="bcast_filter:score_low")],
            [InlineKeyboardButton(text="🟠 Задовільно (35–50)", callback_data="bcast_filter:score_satisf")],
            [InlineKeyboardButton(text="🟡 Середньо (51–80)", callback_data="bcast_filter:score_mid")],
            [InlineKeyboardButton(text="🟢 Добре (81–140)", callback_data="bcast_filter:score_high")],
            [InlineKeyboardButton(text="🎓 2027 рік", callback_data="bcast_filter:course3"),
             InlineKeyboardButton(text="🎓 2026 рік", callback_data="bcast_filter:course4")],
            [InlineKeyboardButton(text="🆕 Ще не починали тест", callback_data="bcast_filter:no_test")],
            [InlineKeyboardButton(text="❌ Скасувати", callback_data="bcast_filter:cancel")],
        ])

    FILTER_LABELS = {
        "all": "👥 Всі користувачі",
        "completed": "✅ Пройшли тест до кінця",
        "not_completed": "⏸ Не завершили тест",
        "score_low": "🔴 Погано (до 35 балів)",
        "score_satisf": "🟠 Задовільно (35–50)",
        "score_mid": "🟡 Середньо (51–80)",
        "score_high": "🟢 Добре (81–140)",
        "course3": "🎓 2027 рік",
        "course4": "🎓 2026 рік",
        "no_test": "🆕 Ще не починали тест",
    }

    async def _get_broadcast_users(filter_key: str) -> list[int]:
        from datetime import datetime, timedelta
        cutoff = datetime.utcnow() - timedelta(days=7)
        async with session_factory() as session:
            if filter_key == "all":
                return list((await session.execute(
                    select(AccessRequest.user_id).where(AccessRequest.approved == True)
                )).scalars().all())

            elif filter_key == "no_test":
                tested = set((await session.execute(
                    select(UserResult.user_id).where(UserResult.created_at >= cutoff).distinct()
                )).scalars().all())
                all_approved = (await session.execute(
                    select(AccessRequest.user_id).where(AccessRequest.approved == True)
                )).scalars().all()
                return [uid for uid in all_approved if uid not in tested]

            elif filter_key == "completed":
                return list((await session.execute(
                    select(UserResult.user_id)
                    .where(UserResult.created_at >= cutoff, UserResult.completed == True)
                    .distinct()
                )).scalars().all())

            elif filter_key == "not_completed":
                return list((await session.execute(
                    select(UserResult.user_id)
                    .where(UserResult.created_at >= cutoff, UserResult.completed == False)
                    .distinct()
                )).scalars().all())

            elif filter_key.startswith("score_"):
                level = filter_key[6:]
                results = (await session.execute(
                    select(UserResult)
                    .where(UserResult.created_at >= cutoff, UserResult.completed == True)
                )).scalars().all()
                uid_set = set()
                for r in results:
                    pct = r.score * 100 // r.total
                    if level == "low" and pct <= 25:
                        uid_set.add(r.user_id)
                    elif level == "satisf" and 26 <= pct <= 36:
                        uid_set.add(r.user_id)
                    elif level == "mid" and 37 <= pct <= 57:
                        uid_set.add(r.user_id)
                    elif level == "high" and pct >= 58:
                        uid_set.add(r.user_id)
                return list(uid_set)

            elif filter_key in ("course3", "course4"):
                course_num = filter_key[-1]
                return list((await session.execute(
                    select(AccessRequest.user_id)
                    .where(
                        AccessRequest.approved == True,
                        AccessRequest.course.like(f"%{course_num}%"),
                    )
                )).scalars().all())
        return []

    async def _start_broadcast(target: Message, state: FSMContext):
        await state.set_state(UserState.broadcast_waiting)
        await target.answer(
            "📣 <b>Розсилка</b>\n\nОбери <b>кому</b> надіслати:",
            parse_mode="HTML",
            reply_markup=broadcast_filter_keyboard(),
        )

    @dp.callback_query(F.data.startswith("bcast_filter:"))
    async def broadcast_filter_chosen(callback: CallbackQuery, state: FSMContext):
        if callback.from_user.id not in ADMIN_IDS:
            await callback.answer("⛔", show_alert=True)
            return
        filter_key = callback.data.split(":")[1]
        await callback.answer()

        if filter_key == "cancel":
            await state.clear()
            await callback.message.edit_reply_markup(reply_markup=None)
            await callback.message.answer("❌ Розсилку скасовано.")
            return

        user_ids = await _get_broadcast_users(filter_key)
        label = FILTER_LABELS.get(filter_key, filter_key)

        if not user_ids:
            await callback.message.answer(f"⚠️ За фільтром «{label}» не знайдено жодного користувача.")
            return

        await state.set_state(UserState.broadcast_filter_set)
        await state.update_data(broadcast_filter=filter_key, broadcast_user_ids=user_ids)
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.message.answer(
            f"✅ Фільтр: <b>{label}</b>\n"
            f"👥 Отримувачів: <b>{len(user_ids)}</b>\n\n"
            "📩 Надішли повідомлення для розсилки\n"
            "(текст, фото, відео, документ — будь-який формат).\n\n"
            "Для скасування: /cancel",
            parse_mode="HTML",
        )

    @dp.message(UserState.broadcast_filter_set)
    async def do_broadcast_filtered(message: Message, state: FSMContext):
        if message.from_user.id not in ADMIN_IDS:
            return
        data = await state.get_data()
        user_ids: list[int] = data.get("broadcast_user_ids", [])
        filter_key = data.get("broadcast_filter", "all")
        label = FILTER_LABELS.get(filter_key, filter_key)
        await state.clear()

        if not user_ids:
            await message.answer("Немає користувачів для розсилки.")
            return

        await message.answer(
            f"⏳ Починаю розсилку для <b>{len(user_ids)}</b> користувачів ({label})...",
            parse_mode="HTML",
        )
        sent, failed = 0, 0
        for user_id in user_ids:
            try:
                await asyncio.sleep(0.05)
                await bot.copy_message(
                    chat_id=user_id,
                    from_chat_id=message.chat.id,
                    message_id=message.message_id,
                )
                sent += 1
            except TelegramRetryAfter as e:
                await asyncio.sleep(e.retry_after + 1)
                try:
                    await bot.copy_message(
                        chat_id=user_id,
                        from_chat_id=message.chat.id,
                        message_id=message.message_id,
                    )
                    sent += 1
                except Exception:
                    failed += 1
            except Exception as e:
                logger.warning("Broadcast failed for %d: %s", user_id, e)
                failed += 1

        await message.answer(
            f"✅ <b>Розсилка завершена!</b>\n\n"
            f"🎯 Фільтр: {label}\n"
            f"📤 Надіслано: <b>{sent}</b>\n"
            f"❌ Помилок: <b>{failed}</b>",
            parse_mode="HTML",
        )

    @dp.message(UserState.broadcast_waiting)
    async def do_broadcast(message: Message, state: FSMContext):
        # Show filter selection again if user sends message without choosing filter
        if message.from_user.id not in ADMIN_IDS:
            return
        await _start_broadcast(message, state)

    # -----------------------------------------------------------------------
    # Access check helper
    # -----------------------------------------------------------------------
    async def check_access(message: Message) -> bool:
        if message.from_user.id in ADMIN_IDS:
            return True
        async with session_factory() as session:
            req = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == message.from_user.id)
            )).scalar()
        if req and req.approved:
            return True
        await message.answer("🔒 Спочатку потрібно отримати доступ. Напиши /start і надішли скріншоти.")
        return False

    # -----------------------------------------------------------------------
    # /stats
    # -----------------------------------------------------------------------
    @dp.message(Command("stats"))
    async def cmd_stats(message: Message):
        if not await check_access(message):
            return
        async with session_factory() as session:
            rows = (await session.execute(
                select(UserResult)
                .where(UserResult.user_id == message.from_user.id)
                .order_by(UserResult.id.desc())
                .limit(5)
            )).scalars().all()
        if not rows:
            await message.answer("Ти ще не проходив тест.")
            return
        lines = ["📊 <b>Твої останні результати:</b>"]
        for r in rows:
            pct = r.score * 100 // r.total
            lines.append(f"• {r.score}/{r.total} ({pct}%)")
        await message.answer("\n".join(lines), parse_mode="HTML")

    # -----------------------------------------------------------------------
    # Quiz start
    # -----------------------------------------------------------------------
    async def start_quiz(message: Message, state: FSMContext, count: int | None = None, user=None, section: str | None = None):
        actual_user = user or message.from_user
        is_admin = actual_user.id in ADMIN_IDS

        async with session_factory() as session:
            req = (await session.execute(
                select(AccessRequest).where(AccessRequest.user_id == actual_user.id)
            )).scalar()
        has_access = (req and req.approved) or is_admin
        if not has_access:
            await message.answer("🔒 Спочатку потрібно отримати доступ. Напиши /start і надішли скріншоти.")
            return

        if req and not req.pib and not is_admin:
            await state.set_state(UserState.filling_pib)
            await message.answer(
                "✏️ Спочатку заповни профіль.\n\nВведи своє <b>ПІБ</b>:",
                parse_mode="HTML",
            )
            return

        async with session_factory() as session:
            query = select(Question).order_by(Question.id)
            if section:
                query = query.where(Question.section == section)
            all_qs = (await session.execute(query)).scalars().all()

        if not all_qs:
            await message.answer("Питання не знайдено.")
            return

        q_ids = [q.id for q in all_qs]
        if count:
            q_ids = random.sample(q_ids, min(count, len(q_ids)))

        section_label = f" — {section}" if section else ""
        await message.answer(
            f"🚀 Починаємо тест{section_label}!\n"
            f"Всього питань: <b>{len(q_ids)}</b>",
            parse_mode="HTML",
        )
        await message.answer(
            "⚠️ <b>ТВІЙ РЕЗУЛЬТАТ І ВІДПОВІДІ</b> будуть доступні <b>24 ГОДИНИ</b>, "
            "після чого <b>АВТОМАТИЧНО ОЧИЩУЮТЬСЯ</b> — але ти завжди зможеш <b>ПРОЙТИ ТЕСТ ЩЕ РАЗ</b>.",
            parse_mode="HTML",
        )

        await state.set_state(UserState.quiz_in_progress)
        await state.set_data({
            "q_ids": q_ids,
            "current": 0,
            "score": 0,
            "user_id": actual_user.id,
            "username": actual_user.username or "",
            "section": section or "",
            "is_admin": is_admin,
        })

        async with session_factory() as session:
            await send_question(bot, message.chat.id, state, session, session_factory)

    @dp.message(Command("quiz"))
    async def cmd_quiz(message: Message, state: FSMContext):
        if not await check_access(message):
            return
        await message.answer("Обери формат тесту:", reply_markup=quiz_menu_keyboard())

    @dp.callback_query(F.data.startswith("startquiz:"))
    async def handle_quiz_choice(callback: CallbackQuery, state: FSMContext):
        mode = callback.data.split(":")[1]
        if mode == "sections":
            await callback.message.edit_reply_markup(reply_markup=None)
            await callback.answer()
            await callback.message.answer("📚 Обери розділ:", reply_markup=sections_keyboard())
            return
        if mode == "back":
            await callback.message.edit_reply_markup(reply_markup=None)
            await callback.answer()
            await callback.message.answer("Обери формат тесту:", reply_markup=quiz_menu_keyboard())
            return
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer()
        count = None if mode == "all" else int(mode)
        await start_quiz(callback.message, state, count=count, user=callback.from_user)

    @dp.callback_query(F.data.startswith("section:"))
    async def handle_section_choice(callback: CallbackQuery, state: FSMContext):
        idx = int(callback.data.split(":")[1])
        section_name = SECTIONS_LIST[idx]
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer()
        await start_quiz(
            callback.message, state,
            section=section_name,
            user=callback.from_user,
        )

    # -----------------------------------------------------------------------
    # Answer handler
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data.startswith("ans:"), UserState.quiz_in_progress)
    async def handle_answer(callback: CallbackQuery, state: FSMContext):
        _, q_id_str, ans_str = callback.data.split(":")
        q_id = int(q_id_str)
        ans_idx = int(ans_str)

        # Cancel the running timer immediately
        cancel_timer(callback.message.chat.id)

        async with session_factory() as session:
            q = (await session.execute(select(Question).where(Question.id == q_id))).scalar_one()

            data = await state.get_data()
            score: int = data["score"]
            current: int = data["current"]
            q_msg_id: int = data.get("q_msg_id", 0)
            timer_msg_id_saved: int = data.get("timer_msg_id", 0)
            chat_id = callback.message.chat.id

            is_correct = ans_idx == q.correct
            chosen_letter = chr(65 + ans_idx)
            correct_letter = chr(65 + q.correct)

            if is_correct:
                score += 1
                feedback = f"✅ <b>Правильно!</b>\n\n{correct_letter}) {q.options[q.correct]}"
            else:
                feedback = (
                    f"❌ <b>Неправильно.</b>\n\n"
                    f"Твоя відповідь: {chosen_letter}) {q.options[ans_idx]}\n\n"
                    f"✅ Правильна відповідь:\n{correct_letter}) {q.options[q.correct]}"
                )

            await callback.answer()

            # Remove answer buttons from question (keep text visible)
            try:
                await callback.message.edit_reply_markup(reply_markup=None)
            except Exception:
                pass

            # Show feedback
            feedback_msg = await safe_send(
                bot.send_message,
                chat_id,
                feedback,
                parse_mode="HTML",
                protect_content=True,
            )

            # Schedule deletion of question, timer and feedback after 24h
            msgs_to_delete = [m for m in [q_msg_id, timer_msg_id_saved,
                                           feedback_msg.message_id if feedback_msg else 0] if m]
            if msgs_to_delete:
                asyncio.create_task(_delete_messages_after(bot, chat_id, msgs_to_delete, delay=86400))

            await state.update_data(score=score, current=current + 1)
            await asyncio.sleep(2.0)
            await send_question(bot, chat_id, state, session, session_factory)

    # -----------------------------------------------------------------------
    # Quiz quit — user presses "🚪 Завершити тест"
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data == "quiz:quit", UserState.quiz_in_progress)
    async def quiz_quit_confirm(callback: CallbackQuery, state: FSMContext):
        await callback.answer()
        data = await state.get_data()
        current = data.get("current", 0)
        total = len(data.get("q_ids", []))

        # Remove answer buttons from current question
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass

        # Cancel timer
        cancel_timer(callback.message.chat.id)

        await state.set_state(UserState.quiz_quit_comment)
        await callback.message.answer(
            f"⚠️ Ти збираєшся завершити тест на питанні <b>{current + 1}/{total}</b>.\n\n"
            "📝 Напиши будь ласка <b>причину</b>, чому вирішив завершити тест раніше часу:",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="↩️ Продовжити тест", callback_data="quiz:resume")],
            ]),
        )

    @dp.callback_query(F.data == "quiz:resume", UserState.quiz_quit_comment)
    async def quiz_resume(callback: CallbackQuery, state: FSMContext):
        await callback.answer()
        await callback.message.edit_reply_markup(reply_markup=None)
        await state.set_state(UserState.quiz_in_progress)
        await callback.message.answer("▶️ Продовжуємо тест!")
        async with session_factory() as session:
            await send_question(bot, callback.message.chat.id, state, session, session_factory)

    @dp.message(UserState.quiz_quit_comment)
    async def quiz_quit_save(message: Message, state: FSMContext):
        # Validate comment — must be non-empty text, at least 5 characters
        if not message.text:
            await message.answer(
                "⚠️ <b>Коментар обов'язковий!</b>\n\n"
                "Будь ласка, напиши текстом причину завершення тесту.\n"
                "Без коментаря перервати тест неможливо.",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="↩️ Продовжити тест", callback_data="quiz:resume")],
                ]),
            )
            return

        comment = message.text.strip()
        if len(comment) < 5:
            await message.answer(
                "⚠️ <b>Коментар занадто короткий!</b>\n\n"
                "Напиши будь ласка детальніше — мінімум 5 символів.\n"
                "Без коментаря перервати тест неможливо.",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="↩️ Продовжити тест", callback_data="quiz:resume")],
                ]),
            )
            return

        data = await state.get_data()
        user_id = data.get("user_id", message.from_user.id)
        username = data.get("username", message.from_user.username or "")
        score = data.get("score", 0)
        current = data.get("current", 0)
        q_ids = data.get("q_ids", [])
        total = len(q_ids)
        section_label = data.get("section", "")

        # Save incomplete result
        async with session_factory() as session:
            result_obj = UserResult(
                user_id=user_id,
                username=username,
                score=score,
                total=total,
                section=section_label,
                stopped_at=current + 1,
                completed=False,
            )
            session.add(result_obj)
            await session.commit()

        await state.clear()
        cancel_timer(message.chat.id)

        await message.answer(
            f"📋 <b>Тест завершено достроково</b>\n\n"
            f"Питань пройдено: <b>{current}/{total}</b>\n"
            f"Правильних відповідей: <b>{score}</b>\n\n"
            f"💬 Твій коментар: <i>{comment}</i>\n\n"
            "Дякуємо за відгук! Якщо захочеш — можеш пройти тест знову 👇",
            parse_mode="HTML",
            reply_markup=quiz_menu_keyboard(),
        )

        # Notify admins
        if ADMIN_IDS:
            for admin_id in ADMIN_IDS:
                try:
                    await bot.send_message(
                        admin_id,
                        f"⚠️ <b>Користувач завершив тест достроково</b>\n\n"
                        f"👤 @{username} (<code>{user_id}</code>)\n"
                        f"📊 Пройдено: {current}/{total} питань\n"
                        f"✅ Правильних: {score}\n"
                        + (f"📚 Розділ: {section_label}\n" if section_label else "")
                        + f"\n💬 Причина: <i>{comment}</i>",
                        parse_mode="HTML",
                        reply_markup=write_to_user_keyboard(user_id),
                    )
                except Exception:
                    pass

    # -----------------------------------------------------------------------
    # Post-test funnel — score range selection
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data.startswith("score_range:"))
    async def handle_score_range(callback: CallbackQuery):
        parts = callback.data.split(":")
        level = parts[1]  # low / satisf / mid / high
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer()

        if level == "low":
            text = (
                "🔴 <b>НИЗЬКИЙ РЕЗУЛЬТАТ — до 35 правильних відповідей</b>\n\n"
                "Якби це був реальний ЄФВВ — <b>ти б не пройшов</b>, тому що ти <b>не набрав пороговий бал</b>. "
                "Проблема може бути в тому, що ти <b>не розумієш логіку тесту</b> та маєш низький рівень. "
                "Цей тест, до речі, побудований за тією ж системою, за якою я веду підготовку. "
                "Якщо тобі цікаво — <b>давай покажу, як це виправити</b>.\n\n"
                "В залежності від того, в якому році тобі складати ЄФВВ, маю <b>два варіанти порад</b>:"
            )

        elif level == "satisf":
            text = (
                "🟠 <b>ЗАДОВІЛЬНИЙ РЕЗУЛЬТАТ — 35–50 правильних відповідей</b>\n\n"
                "Ви зараз у <b>найнебезпечнішій точці підготовки</b>. Здається, що результат ніби вже «нормальний»: "
                "<b>пороговий набрано</b> і навіть трохи більше. Але саме тут криється <b>найбільший ризик</b>.\n\n"
                "На реальному ЄФВВ результат може бути <b>менше порогового</b> через:\n"
                "• специфічні або нетипові питання;\n"
                "• стрес і хвилювання;\n"
                "• невпевненість у базових темах.\n\n"
                "В залежності від того, в якому році тобі складати ЄФВВ, маю <b>два варіанти порад</b>:"
            )

        elif level == "mid":
            text = (
                "🟡 <b>СЕРЕДНІЙ РЕЗУЛЬТАТ — 51–80 правильних відповідей</b>\n\n"
                "Ви зараз у <b>найнебезпечнішій точці</b>: здається, що <b>«майже готові»</b>.\n"
                "Але на ЄФВВ не працює <b>«плюс-мінус»</b> — працює тільки <b>система</b>.\n\n"
                "Моя точка зору:\n"
                "• <b>50–70</b> — хороший результат, але <b>НЕ бюджет</b>;\n"
                "• <b>70–80+</b> — при правильній підготовці можна набирати стабільно <b>80+</b>, а це вже рівень для бюджету.\n\n"
                "Різниця — <b>не в обсязі вивченого</b>, а в <b>системній підготовці</b>.\n\n"
                "В залежності від того, на якому ти курсі, маю <b>два варіанти порад</b>:"
            )

        else:  # high
            text = (
                "🟢 <b>ВИСОКИЙ РЕЗУЛЬТАТ — 80–140 правильних відповідей</b>\n\n"
                "Результат справді <b>сильний</b>. Але ключове питання — чи зможете ви <b>повторити його на реальному іспиті</b>?\n\n"
                "Тому що ЄФВВ — це не лише про знання, а про <b>готовність до формату іспиту в стресових умовах</b>, "
                "де <b>кожен бал на вагу золота</b>. І дуже часто сильні студенти через невпевненість втрачають бали.\n\n"
                "В залежності від того, на якому ти курсі, маю <b>два варіанти порад</b>:"
            )
        await callback.message.answer(text, parse_mode="HTML", reply_markup=course_keyboard(level))

    # -----------------------------------------------------------------------
    # Funnel — course selection (3 or 4)
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data.startswith("funnel:course"))
    async def handle_course_choice(callback: CallbackQuery):
        parts = callback.data.split(":")
        course = parts[1]   # course3 or course4
        level = parts[2] if len(parts) > 2 else ""
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer()

        if course == "course4":
            await callback.message.answer(
                "Тобі зараз особливо важливо не втрачати час і сфокусуватися на найважливішому.\n\n"
                "Обов'язково переглянь вебінар із повним переліком ТОП-тем, які потрібно повторити перед іспитом, "
                "а також порадами, які варто врахувати безпосередньо на ЄФВВ:\n\n"
                f"👉 <a href='{WEBINAR_URL}'>ПОСИЛАННЯ НА ВЕБІНАР</a>\n\n"
                f"Також заходь в <a href='{INSTAGRAM_URL}'>Instagram</a> — там я регулярно публікую:\n"
                "• ТОП-теми до ЄФВВ;\n"
                "• найпоширеніші помилки;\n"
                "• перелік тем, які обов'язково треба повторити перед іспитом.\n\n"
                "Слідкуй, повторюй і добирай ті бали, яких може не вистачити саме на іспиті.\n\n"
                "А якщо надалі буде потрібна допомога з підготовкою до більш складного випускного іспиту в магістратурі — ЄДКІ, "
                "шукай анкету в шапці профілю.\n\n"
                "Орієнтовний старт набору на курс ЄДКІ 2027 — вересень–жовтень 2026 року. "
                "Найвигідніші умови традиційно отримують ті, хто приходить першими.\n\n"
                "З Мамою ЄФВВ та ЄДКІ складають всі! 💙",
                parse_mode="HTML",
                reply_markup=restart_keyboard(),
            )
        else:  # course3 — text depends on score level
            if level == "low":
                text = (
                    "🔶 Дуже добре, що ви пройшли симуляцію ЄФВВ <b>ЗАЗДАЛЕГІДЬ</b>.\n\n"
                    "Зараз ви вже бачите іспит <b>ЗСЕРЕДИНИ</b> і через свій результат розумієте <b>НЕОБХІДНІСТЬ ПІДГОТОВКИ</b>.\n\n"
                    "🤎 Поточний результат — це не той рівень, на який варто орієнтуватися, але це <b>НЕ КРИТИЧНО, ЯКЩО ВЧАСНО ПОЧАТИ ПІДГОТОВКУ</b>.\n\n"
                    "🔶 Обсяг матеріалу великий, і результат потрібно <b>ПІДНІМАТИ ПО ВСІХ БЛОКАХ</b>, тому важливо не відкладати.\n\n"
                    "🔶 ЄФВВ — це не лише знання, а й <b>ОСОБЛИВА ПІДГОТОВКА ДО СПЕЦІАЛЬНОГО ФОРМАТУ ІСПИТУ</b>.\n"
                    "Без цього <b>ВАШ РЕЗУЛЬТАТ</b>, як правило, <b>ЗАЛИШИТЬСЯ НА ТОМУ Ж РІВНІ</b>.\n\n"
                    "👉 Якщо хочете впевнено здолати свій поріг і підвищити результат — напишіть мені, і я розповім деталі підготовки та запропоную <b>НАЙВИГІДНІШІ УМОВИ</b>."
                )
            elif level == "satisf":
                text = (
                    "🔶 Дуже добре, що ви пройшли симуляцію ЄФВВ <b>ЗАЗДАЛЕГІДЬ</b>.\n\n"
                    "Це вже ваша <b>ПЕРЕВАГА</b>: ви побачили іспит <b>ЗСЕРЕДИНИ</b> і зрозуміли, "
                    "що підготовка має бути <b>СИСТЕМНОЮ</b>.\n\n"
                    "🤎 Так, поточний результат — ще не той, на який варто орієнтуватися. "
                    "Але у вас є <b>ГОЛОВНЕ</b> — <b>ДОСТАТНЬО ЧАСУ</b>, щоб усе виправити.\n\n"
                    "Більшість не показує високий результат з першої спроби, тому що ЄФВВ — це не лише знання, "
                    "а й <b>УМІННЯ ПРАЦЮВАТИ У ФОРМАТІ ІСПИТУ</b>.\n\n"
                    "🔶 Тому зараз найкраще рішення — почати підготовку <b>ЗАВЧАСНО</b> і <b>ПРАВИЛЬНО</b>.\n\n"
                    "👉 Якщо вас зацікавила підготовка зі мною — розповім деталі та запропоную <b>НАЙВИГІДНІШІ УМОВИ</b>. 🤎"
                )
            elif level == "mid":
                text = (
                    "🔶 Дуже добре, що ви пройшли симуляцію ЄФВВ <b>ЗАЗДАЛЕГІДЬ</b>.\n\n"
                    "Ваш результат вже <b>ВИЩЕ СЕРЕДНЬОГО</b> — і це показник, що у вас є <b>СИЛЬНА БАЗА</b>.\n\n"
                    "🤎 Ви вже бачите іспит <b>ЗСЕРЕДИНИ</b> і розумієте його логіку.\n"
                    "Але важливо врахувати: на реальному іспиті результат може <b>ПРОСІДАТИ</b>, "
                    "якщо немає чітко відпрацьованої стратегії роботи з тестом (час, пастки, формулювання).\n\n"
                    "🔶 Тепер ключове — <b>ДОТИСНУТИ РЕЗУЛЬТАТ ДО ВИСОКОГО РІВНЯ</b> і зробити його <b>СТАБІЛЬНИМ</b>, "
                    "тому що ЄФВВ — це не лише знання, а й <b>ТОЧНА РОБОТА У ФОРМАТІ ІСПИТУ</b>.\n\n"
                    "🔶 Саме зараз найкращий момент, щоб системно допрацювати слабкі місця і вийти на <b>МАКСИМУМ</b>.\n\n"
                    "👉 Якщо хочете підсилити результат і впевнено показати його на іспиті — "
                    "розповім деталі підготовки та запропоную <b>НАЙВИГІДНІШІ УМОВИ</b>. 🤎"
                )
            else:  # high
                text = (
                    "🔶 Дуже добре, що ви пройшли симуляцію ЄФВВ <b>ЗАЗДАЛЕГІДЬ</b>.\n\n"
                    "Це ваша <b>ПЕРЕВАГА</b>: ви вже бачите іспит <b>ЗСЕРЕДИНИ</b> і розумієте його логіку.\n\n"
                    "🤎 Ваш результат уже <b>ВИСОКИЙ</b>, але важливо інше — "
                    "<b>ЧИ ЗМОЖЕТЕ ВИ ЙОГО ПОВТОРИТИ НА ІСПИТІ</b>.\n\n"
                    "Без системної підготовки навіть сильні студенти <b>ВТРАЧАЮТЬ БАЛИ</b> "
                    "через формат, час і неочевидні пастки тесту.\n\n"
                    "🔶 Зараз ключове — <b>ЗАКРІПИТИ РЕЗУЛЬТАТ</b> і <b>ДОТИСНУТИ ЙОГО ДО МАКСИМУМУ</b>.\n"
                    "Саме тут і починається <b>СПЕЦИФІЧНА ПІДГОТОВКА ДЛЯ ВІДМІННИКІВ</b>.\n\n"
                    "👉 Якщо хочете не просто «добре скласти», а <b>ГАРАНТОВАНО УТРИМАТИ ВИСОКИЙ РІВЕНЬ</b> — "
                    "напишіть, розповім деталі та запропоную <b>НАЙКРАЩІ УМОВИ</b>. 🤎"
                )

            await callback.message.answer(
                text,
                parse_mode="HTML",
                reply_markup=interested_keyboard(),
            )
            asyncio.create_task(_send_warmup(bot, callback.from_user.id, delay_hours=6))

    # -----------------------------------------------------------------------
    # Funnel — "Так, цікаво" / "Хочу дізнатися більше"
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data.startswith("funnel:interested") | F.data.startswith("funnel:want_more"))
    async def handle_interested(callback: CallbackQuery):
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer()
        await callback.message.answer(
            "<b>ФОРМУЮ РАННІЙ СПИСОК (ЛИСТ ОЧІКУВАННЯ) НА 2027 РІК</b>\n\n"
            "Беру обмежену кількість студентів, адже працюю з кожним <b>ІНДИВІДУАЛЬНО</b> — "
            "персональний підхід, постійний зворотний зв'язок і повний супровід у підготовці.\n\n"
            "Хочете дізнатися, чи підійде вам цей формат і чи встигнете <b>ЯКІСНО ПІДГОТУВАТИСЯ</b>?\n\n"
            "Заповнюйте анкету — і я <b>ЗВʼЯЖУСЬ З ВАМИ В ПЕРШУ ЧЕРГУ</b> та запропоную "
            "<b>НАЙНИЖЧУ ВАРТІСТЬ</b> на <b>КУРС ЄФВВ 2027</b> + <b>МАКСИМУМ БОНУСІВ</b>.",
            parse_mode="HTML",
            reply_markup=fill_form_keyboard(),
        )

    async def _send_warmup(bot: Bot, user_id: int, delay_hours: float = 6):
        """Send warm-up message after delay if user hasn't responded."""
        await asyncio.sleep(delay_hours * 3600)
        try:
            await bot.send_message(
                user_id,
                "70% студентів, які проходять СИМУЛЯЦІЮ думають, що «ще є час», "
                "а потім пишуть мені перед іспитом, як можна наздогнати все пропущене?\n\n"
                "Подивіться, як виглядає підготовка у мене:\n"
                "— без зайвих конспектів і води\n"
                "— тільки те, що реально дає бали на іспиті\n"
                "— тренування у форматі тестів ЄФВВ\n"
                "— системна підготовка через мої індивідуальні консультації та моя «рука на пульсі» щодо вашого прогресу\n\n"
                "👉 Історична довідка: у мене всі студенти доходять до іспиту і складають з першого разу. "
                "Навіть ті, хто починав майже з нуля. А якщо докладають максимум зусиль, то отримують бюджет, "
                "навіть ті, хто на це не розраховував.",
                parse_mode="HTML",
                reply_markup=want_more_keyboard(),
            )
        except Exception as e:
            logger.warning("Warmup send failed for %d: %s", user_id, e)

    # -----------------------------------------------------------------------
    # Restart quiz callback
    # -----------------------------------------------------------------------
    @dp.callback_query(F.data == "restart_quiz")
    async def handle_restart(callback: CallbackQuery, state: FSMContext):
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.answer()
        await callback.message.answer(
            "🔄 Обери формат тесту:",
            reply_markup=quiz_menu_keyboard(),
        )

    # -----------------------------------------------------------------------
    # Background cleanup task — delete results older than 7 days
    # -----------------------------------------------------------------------
    async def auto_cleanup():
        from datetime import datetime, timedelta
        from sqlalchemy import delete as sa_delete
        while True:
            try:
                cutoff = datetime.utcnow() - timedelta(days=7)
                async with session_factory() as session:
                    result = await session.execute(
                        sa_delete(UserResult).where(UserResult.created_at < cutoff)
                    )
                    deleted = result.rowcount
                    await session.commit()
                    if deleted:
                        logger.info("Auto-cleanup: deleted %d old results", deleted)
            except Exception as e:
                logger.warning("Auto-cleanup error: %s", e)
            await asyncio.sleep(6 * 3600)  # run every 6 hours

    asyncio.create_task(auto_cleanup())

    # -----------------------------------------------------------------------
    # Set bot commands (shows "Start" button in empty chats)
    # -----------------------------------------------------------------------
    from aiogram.types import BotCommand
    await bot.set_my_commands([
        BotCommand(command="start", description="🚀 Розпочати / Головне меню"),
        BotCommand(command="quiz", description="📝 Пройти тест"),
        BotCommand(command="stats", description="📊 Мої результати"),
        BotCommand(command="cancel", description="❌ Скасувати поточну дію"),
    ])

    # -----------------------------------------------------------------------
    # Webhook or polling
    # -----------------------------------------------------------------------
    if WEBHOOK_DOMAIN:
        from aiohttp import web
        from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application

        webhook_url = f"{WEBHOOK_DOMAIN}{WEBHOOK_PATH}"
        await bot.set_webhook(
            url=webhook_url,
            secret_token=WEBHOOK_SECRET,
            drop_pending_updates=True,
        )
        logger.info("Webhook set: %s", webhook_url)

        app = web.Application()

        SimpleRequestHandler(dispatcher=dp, bot=bot, secret_token=WEBHOOK_SECRET).register(app, path=WEBHOOK_PATH)
        setup_application(app, dp, bot=bot)

        runner = web.AppRunner(app)
        await runner.setup()
        site = web.TCPSite(runner, host="0.0.0.0", port=8000)
        await site.start()
        logger.info("Webhook server started on :8000")

        await asyncio.Event().wait()
    else:
        await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
